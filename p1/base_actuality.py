
from openpyxl import load_workbook

import psycopg2
import re


orderBaseFile = 'C:\\microservices\\order_base.xlsm'
dlWithoutGen23BaseFile = 'C:\\microservices\\dl_without_gen23.xlsx'
organizations = []
dls = []



class Organization:

    def __init__(self, docid,
                 gid,
                 inn,
                 full_organization_name,
                 organization_name,
                 services,
                 doc_num,
                 doc_date,
                 order_start_date,
                 order_end_date,
                 order_status,
                 order_activity,
                 document_price,
                 document_price_with_nds,
                 doc_prepare_date,
                 doc_eosdo_date,
                 doc_sending_date,
                 order_activity_date,
                 doc_eosdo_num,
                 doc_eosdo_link,
                 ):
        self.docid = docid
        self.gid = gid
        self.inn = inn
        self.full_organization_name = full_organization_name
        self.organization_name = organization_name
        self.services = services
        self.doc_num = doc_num
        self.doc_date = doc_date
        self.order_start_date = order_start_date
        self.order_end_date = order_end_date
        self.order_status = order_status
        self.order_activity = order_activity
        self.document_price = document_price
        self.document_price_with_nds = document_price_with_nds
        self.doc_prepare_date = doc_prepare_date
        self.doc_eosdo_date = doc_eosdo_date
        self.doc_sending_date = doc_sending_date
        self.order_activity_date = order_activity_date
        self.doc_eosdo_num = doc_eosdo_num
        self.doc_eosdo_link = doc_eosdo_link

    def set_city(self, city):
        self._city = city

    def set_filial_name(self, filial_name):
        self._filial_name = filial_name

    def set_order_num(self, order_num):
        self._order_num = order_num

    def set_order_date(self, order_date):
        self._order_date = order_date

    def get_gid(self):
        return self.gid

    def get_docid(self):
        return self.docid

    def get_doc_num(self):
        return self.doc_num

    def get_doc_date(self):
        return self.doc_date

    def get_inn(self):
        return self.inn

    def get_full_organization_name(self):
        return self.full_organization_name

    def get_organization_name(self):
        return self.organization_name

    def get_services(self):
        return self.services

    def get_order_start_date(self):
        return self.order_start_date

    def get_order_end_date(self):
        return self.order_end_date

    def get_order_status(self):
        return self.order_status

    def get_order_activity(self):
        return self.order_activity

    def get_document_price(self):
        return self.document_price

    def get_document_price_with_nds(self):
        return self.document_price_with_nds

    def get_doc_prepare_date(self):
        return self.doc_prepare_date

    def get_doc_eosdo_date(self):
        return self.doc_eosdo_date

    def get_doc_sending_date(self):
        return self.doc_sending_date

    def get_order_activity_date(self):
        return self.order_activity_date

    def get_doc_eosdo_num(self):
        return self.doc_eosdo_num

    def get_doc_eosdo_link(self):
        return self.doc_eosdo_link

class Dl:

    def __init__(self, gid, filialName, city, order_num, order_date, fio, dl_name, work_tel, mob_tel,e_mail):
        self.gid = gid
        self.filialName = filialName
        self.city = city
        self.order_num = order_num
        self.order_date = order_date
        self.fio = fio
        self.dl_name = dl_name
        self.work_tel = work_tel
        self.mob_tel = mob_tel
        self.e_mail = e_mail

    def get_filialName(self):
            return self.filialName

    def get_mob_tel(self):
            return self.mob_tel

    def get_work_tel(self):
            return self.work_tel

    def get_order_num(self):
            return self.order_num

    def get_gid(self):
            return self.gid

    def get_city(self):
            return self.city

    def get_fio(self):
            return self.fio

    def get_e_mail(self):
            return self.e_mail

    def get_order_date(self):
            return self.order_date

    def get_dl_name(self):
            return self.dl_name

    def __repr__(self):
        return str(self.__dict__)


def readOrderBase():

    wb = load_workbook(orderBaseFile, data_only=True)

    ws = wb['Договоры']
    j = 1
    for i in range (1, ws.max_row):
        cell = ws["A" + str(i)]
        if "ID документа" in str(cell.value):
            break
        else:
            j += 1

    j += 1

    for row in range(j, ws.max_row):

        docid = ws["A" + str(row)].value
        gid = ws["F" + str(row)].value
        inn = ws["G" + str(row)].value
        full_organization_name = ws["I" + str(row)].value
        organization_name = ws["J" + str(row)].value
        services = ws["Q" + str(row)].value
        doc_num = ws["L" + str(row)].value

        if str(ws["M" + str(row)].value) == "None" or isinstance(ws["M" + str(row)].value, str):
            doc_date = '2030-01-01 00:00:00'
        else:
            doc_date = ws["M" + str(row)].value


        if str(ws["R" + str(row)].value) == "None" or isinstance(ws["R" + str(row)].value, str):
            order_start_date = '2030-01-01 00:00:00'
        else:
            order_start_date = ws["R" + str(row)].value

        if str(ws["S" + str(row)].value) == "None" or isinstance(ws["S" + str(row)].value, str):
            order_end_date = '2030-01-01 00:00:00'
        else:
            order_end_date = ws["S" + str(row)].value

        order_status = ws["U" + str(row)].value
        order_activity = ws["V" + str(row)].value
        document_price = ws["AD" + str(row)].value
        document_price_with_nds = ws["AE" + str(row)].value
        if str(ws["AG" + str(row)].value) == "None" or isinstance(ws["AG" + str(row)].value, str):
            doc_prepare_date = '2030-01-01 00:00:00'
        else:
            doc_prepare_date = ws["AG" + str(row)].value

        if str(ws["AH" + str(row)].value) == "None" or isinstance(ws["AH" + str(row)].value, str):
            doc_eosdo_date = '2030-01-01 00:00:00'
        else:
            doc_eosdo_date = ws["AH" + str(row)].value

        if str(ws["AI" + str(row)].value) == "None" or isinstance(ws["AI" + str(row)].value, str):
            doc_sending_date = '2030-01-01 00:00:00'
        else:
            doc_sending_date = ws["AI" + str(row)].value

        if str(ws["AI" + str(row)].value) == "None" or isinstance(ws["AI" + str(row)].value, str):
            order_activity_date = '2030-01-01 00:00:00'
        else:
            order_activity_date = ws["AI" + str(row)].value
        doc_eosdo_num = ws["AO" + str(row)].value
        try:
            doc_eosdo_link = ws["AP" + str(row)].hyperlink.target
        except AttributeError:
            doc_eosdo_link = ""

        if str(ws["E" + str(row)].value) == "ЦКЗ":
            newOrganization = Organization(docid, gid, inn, full_organization_name, organization_name, services, doc_num, doc_date, order_start_date,
                     order_end_date,
                     order_status,
                     order_activity,
                     document_price,
                     document_price_with_nds,
                     doc_prepare_date,
                     doc_eosdo_date,
                     doc_sending_date,
                     order_activity_date,
                     doc_eosdo_num,
                     doc_eosdo_link)

            organizations.append(newOrganization)


def readDlBase():

        wb = load_workbook(dlWithoutGen23BaseFile)
        ws = wb.get_sheet_by_name('Лист1')
        for row in range(2, ws.max_row):
            if ws["F" + str(row)].value:
                gid = ws["B" + str(row)].value
                filialName = ws["D" + str(row)].value
                city = ws["E" + str(row)].value
                order_num = ws["F" + str(row)].value
                order_date = ws["G" + str(row)].value
                dl_name = ws["H" + str(row)].value
                fio = ws["I" + str(row)].value
                work_tel = ws["J" + str(row)].value
                mob_tel = ws["K" + str(row)].value
                e_mail = ws["L" + str(row)].value

                newDl = Dl(gid,filialName, city, order_num, order_date, fio, dl_name, work_tel, mob_tel, e_mail)

                dls.append(newDl)

def fetchOrgDB():


    for i in organizations:

        conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                password='postgres', host='localhost')
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM administrators_organizations WHERE docid = \'{1}\' AND gid = \'{0}\' AND inn = \'{2}\''.format(
            i.get_gid(),
            i.get_docid(),
            i.get_inn()
        ))
        conn.commit()

        orgRecords = cursor.fetchall()

        cursor.close()
        conn.close()

        if len(orgRecords) == 0:


            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM administrators_orderstatuses WHERE '
                           'status_name = \'{0}\''.format(i.get_order_status()))
            conn.commit()

            records = cursor.fetchall()
            cursor.close()
            conn.close()

            idorderstatus = re.sub("[^0-9]", "", str(records[0]))

            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM administrators_orderactivities WHERE '
                           'activity_name = \'{0}\''.format(i.get_order_activity()))

            records = cursor.fetchall()
            conn.commit()

            cursor.close()
            conn.close()

            try:
                idorderactivity = re.sub("[^0-9]", "", str(records[0]))
            except IndexError as error:
                idorderactivity = 1

            try:
                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
                cursor = conn.cursor()
                sql = 'SELECT MAX(id) FROM administrators_organizations'


                cursor.execute(sql)
                conn.commit()



                sql = 'SELECT MAX(id) FROM administrators_organizations'

                try:
                    cursor.execute(sql)
                    conn.commit()

                    lastId = cursor.fetchall()
                    lastId = int(re.sub("[^0-9]", "", str(lastId)))
                    lastId += 1
                except Exception as e:
                    lastId = 1

                sql = 'INSERT INTO public.administrators_organizations(id, docid, gid, inn, full_organization_name, organization_name, services, doc_num, doc_date, order_start_date, order_end_date, document_price, document_price_with_nds, doc_prepare_date, doc_eosdo_date, doc_sending_date, order_activity_date, doc_eosdo_num, doc_eosdo_link, order_activity_id, order_status_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', \'{4}\', \'{5}\', \'{6}\', \'{7}\', \'{8}\', \'{9}\', \'{10}\', \'{11}\', \'{12}\', \'{13}\', \'{14}\', \'{15}\', \'{16}\', \'{17}\', \'{18}\', {19}, {20})'.format(
                    lastId,
                    i.get_docid(),
                    i.get_gid(),
                    i.get_inn(),
                    i.get_full_organization_name(),
                    i.get_organization_name(),
                    i.get_services(),
                    i.get_doc_num(),
                    i.get_doc_date(),
                    i.get_order_start_date(),
                    i.get_order_end_date(),
                    i.get_document_price(),
                    i.get_document_price_with_nds(),
                    i.get_doc_prepare_date(),
                    i.get_doc_eosdo_date(),
                    i.get_doc_sending_date(),
                    i.get_order_activity_date(),
                    i.get_doc_eosdo_num(),
                    i.get_doc_eosdo_link(),
                    idorderactivity,
                    idorderstatus
                )

                print(sql)

                cursor.execute(sql)
                conn.commit()

                cursor.close()

            except Exception as error:
                print("Failed to insert record into Laptop table {}".format(error))

            finally:
                conn.close()
                print("MySQL connection is closed")

        else:
            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM administrators_orderstatuses WHERE '
                           'status_name = \'{0}\''.format(i.get_order_status()))
            conn.commit()

            records = cursor.fetchall()
            cursor.close()
            conn.close()

            idorderstatus = re.sub("[^0-9]", "", str(records[0]))

            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()

            cursor.execute('SELECT id FROM administrators_orderactivities WHERE '
                           'activity_name = \'{0}\''.format(i.get_order_activity()))

            records = cursor.fetchall()
            conn.commit()

            cursor.close()
            conn.close()

            try:
                idorderactivity = re.sub("[^0-9]", "", str(records[0]))
            except IndexError as error:
                idorderactivity = 1

            try:
                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
                cursor = conn.cursor()

                orgId = orgRecords[0][0]

                sql = 'DELETE FROM administrators_order_admin_ib WHERE 	order_organization_id = {0}'.format(
                    orgId)
                print(sql)
                try:
                    cursor.execute(sql)
                    conn.commit()


                except Exception as e:
                    print("Failed to insert record into Laptop table {}".format(error))

                sql = 'DELETE FROM administrators_organizations WHERE id = {0}'.format(orgId)
                print(sql)
                try:
                    cursor.execute(sql)
                    conn.commit()


                except Exception as e:
                    print("Failed to insert record into Laptop table {}".format(error))



                sql = 'INSERT INTO public.administrators_organizations(id, docid, gid, inn, full_organization_name, organization_name, services, doc_num, doc_date, order_start_date, order_end_date, document_price, document_price_with_nds, doc_prepare_date, doc_eosdo_date, doc_sending_date, order_activity_date, doc_eosdo_num, doc_eosdo_link, order_activity_id, order_status_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', \'{4}\', \'{5}\', \'{6}\', \'{7}\', \'{8}\', \'{9}\', \'{10}\', \'{11}\', \'{12}\', \'{13}\', \'{14}\', \'{15}\', \'{16}\', \'{17}\', \'{18}\', {19}, {20})'.format(
                    orgId,
                    i.get_docid(),
                    i.get_gid(),
                    i.get_inn(),
                    i.get_full_organization_name(),
                    i.get_organization_name(),
                    i.get_services(),
                    i.get_doc_num(),
                    i.get_doc_date(),
                    i.get_order_start_date(),
                    i.get_order_end_date(),
                    i.get_document_price(),
                    i.get_document_price_with_nds(),
                    i.get_doc_prepare_date(),
                    i.get_doc_eosdo_date(),
                    i.get_doc_sending_date(),
                    i.get_order_activity_date(),
                    i.get_doc_eosdo_num(),
                    i.get_doc_eosdo_link(),
                    idorderactivity,
                    idorderstatus
                )

                print(sql)

                cursor.execute(sql)
                conn.commit()

                cursor.close()

            except Exception as error:
                print("Failed to insert record into Laptop table {}".format(error))

            finally:
                conn.close()
                print("MySQL connection is closed")



def fetchDlDB():


    for i in dls:


        if str(i.get_gid()) != "None":
            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()

            cursor.execute(
                'SELECT * FROM administrators_administrators WHERE fio = \'{0}\' AND e_mail = \'{1}\''.format(
                    i.get_fio(),
                    i.get_e_mail()
                ))
            conn.commit()

            adminRrecords = cursor.fetchall()

            cursor.close()
            conn.close()


            if len(adminRrecords) == 0:
                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
                cursor = conn.cursor()

                cursor.execute('SELECT id FROM administrators_dlnames WHERE '
                               '		dl_name = \'{0}\''.format(i.get_dl_name()))

                records = cursor.fetchall()
                conn.commit()

                cursor.close()
                conn.close()

                try:
                    dl_name_id = re.sub("[^0-9]", "", str(records[0]))
                except IndexError as error:
                    dl_name_id = 1

                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
                cursor = conn.cursor()

                cursor.execute('SELECT id FROM administrators_organizations WHERE '
                               '			gid = \'{0}\''.format(i.get_gid()))

                records = cursor.fetchall()
                conn.commit()

                cursor.close()
                conn.close()
                for record in records:

                    try:
                        conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                password='postgres', host='localhost')
                        cursor = conn.cursor()

                        cursor.execute(
                            'SELECT * FROM administrators_administrators WHERE fio = \'{0}\' AND e_mail = \'{1}\''.format(
                                i.get_fio(),
                                i.get_e_mail()
                            ))
                        conn.commit()

                        adminRrecords = cursor.fetchall()

                        cursor.close()
                        conn.close()
                        if len(adminRrecords) == 0:
                            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                    password='postgres', host='localhost')
                            cursor = conn.cursor()

                            sql = 'SELECT MAX(id) FROM administrators_administrators'


                            try:
                                cursor.execute(sql)
                                conn.commit()

                                lastDLId = cursor.fetchall()
                                lastDLId = int(re.sub("[^0-9]", "", str(lastDLId)))
                                lastDLId += 1
                            except Exception as e:
                                lastDLId = 1
                            sql = 'INSERT INTO public.administrators_administrators(id, fio, e_mail, work_telephone, mobile_telephone, city, filial_name, dl_name_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', \'{4}\', \'{5}\', \'{6}\', {7})'.format(
                                lastDLId,
                                i.get_fio(),
                                i.get_e_mail(),
                                i.get_work_tel(),
                                i.get_mob_tel(),
                                i.get_city(),
                                i.get_filialName(),
                                dl_name_id
                            )

                            print(sql)

                            cursor.execute(sql)
                            conn.commit()

                            cursor.close()



                    except Exception as error:
                        print("Failed to insert record into Laptop table {}".format(error))


                    finally:
                        conn.close()
                        print("Success")

                        try:
                            org_id = re.sub("[^0-9]", "", str(record))
                        except IndexError as error:
                            org_id = 0

                        try:

                            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                    password='postgres', host='localhost')
                            cursor = conn.cursor()
                            sql = 'SELECT * FROM administrators_order_admin_ib WHERE order_num = \'{0}\' AND order_date = \'{1}\''.format(
                                i.get_order_num(),
                                i.get_order_date()
                            )

                            cursor.execute(sql)
                            conn.commit()

                            orderRrecords = cursor.fetchall()

                            if len(orderRrecords) == 0:
                                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                        password='postgres', host='localhost')
                                cursor = conn.cursor()
                                sql = 'SELECT MAX(id) FROM administrators_order_admin_ib'


                                try:
                                    cursor.execute(sql)
                                    conn.commit()

                                    lastOrgId = cursor.fetchall()
                                    lastOrgId = int(re.sub("[^0-9]", "", str(lastOrgId)))
                                    lastOrgId += 1
                                except Exception as e:
                                    lastOrgId = 1
                                sql = 'INSERT INTO public.administrators_order_admin_ib(id, order_num, order_date, order_file, order_administrator_id, order_organization_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', {4}, {5})'.format(
                                    lastOrgId,
                                    i.get_order_num(),
                                    i.get_order_date(),
                                    'mainapp\\admin_ib_orders\\' + str(i.get_gid()) + '.pdf',
                                    lastDLId,
                                    org_id
                                )

                                print(sql)

                                cursor.execute(sql)
                                conn.commit()

                                cursor.close()


                        except Exception as error:
                            print("Failed to insert record into Laptop table {}".format(error))

                        finally:
                            conn.close()
                            print("Success")
            else:
                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
                cursor = conn.cursor()

                cursor.execute('SELECT id FROM administrators_dlnames WHERE '
                               '		dl_name = \'{0}\''.format(i.get_dl_name()))

                records = cursor.fetchall()
                conn.commit()

                cursor.close()
                conn.close()

                try:
                    dl_name_id = re.sub("[^0-9]", "", str(records[0]))
                except IndexError as error:
                    dl_name_id = 1

                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
                cursor = conn.cursor()

                cursor.execute('SELECT id FROM administrators_organizations WHERE '
                               '			gid = \'{0}\''.format(i.get_gid()))

                records = cursor.fetchall()
                conn.commit()

                cursor.close()
                conn.close()
                for record in records:

                    try:
                        conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                password='postgres', host='localhost')
                        cursor = conn.cursor()

                        cursor.execute(
                            'SELECT * FROM administrators_administrators WHERE fio = \'{0}\' AND e_mail = \'{1}\''.format(
                                i.get_fio(),
                                i.get_e_mail()
                            ))
                        conn.commit()

                        adminRrecords = cursor.fetchall()

                        cursor.close()
                        conn.close()
                        if len(adminRrecords) == 0:
                            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                    password='postgres', host='localhost')
                            cursor = conn.cursor()

                            sql = 'SELECT MAX(id) FROM administrators_administrators'

                            try:
                                cursor.execute(sql)
                                conn.commit()

                                lastDLId = cursor.fetchall()
                                lastDLId = int(re.sub("[^0-9]", "", str(lastDLId)))
                                lastDLId += 1
                            except Exception as e:
                                lastDLId = 1
                            sql = 'INSERT INTO public.administrators_administrators(id, fio, e_mail, work_telephone, mobile_telephone, city, filial_name, dl_name_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', \'{4}\', \'{5}\', \'{6}\', {7})'.format(
                                lastDLId,
                                i.get_fio(),
                                i.get_e_mail(),
                                i.get_work_tel(),
                                i.get_mob_tel(),
                                i.get_city(),
                                i.get_filialName(),
                                dl_name_id
                            )

                            print(sql)

                            cursor.execute(sql)
                            conn.commit()

                            cursor.close()
                        else:
                            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                    password='postgres', host='localhost')
                            cursor = conn.cursor()

                            adminId = adminRrecords[0][0]

                            sql = 'DELETE FROM administrators_order_admin_ib WHERE order_administrator_id = {0}'.format(
                                adminId)

                            print(sql)
                            try:
                                cursor.execute(sql)
                                conn.commit()
                            except Exception as e:
                                print("Failed to delete record into Laptop table {}".format(error))

                            sql = 'DELETE FROM administrators_administrators WHERE id = {0}'.format(adminId)
                            print(sql)

                            try:
                                cursor.execute(sql)
                                conn.commit()
                            except Exception as e:
                                print("Failed to delete record into Laptop table {}".format(error))



                            except Exception as e:
                                print("Failed to delete record into Laptop table {}".format(error))

                            sql = 'INSERT INTO public.administrators_administrators(id, fio, e_mail, work_telephone, mobile_telephone, city, filial_name, dl_name_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', \'{4}\', \'{5}\', \'{6}\', {7})'.format(
                                adminId,
                                i.get_fio(),
                                i.get_e_mail(),
                                i.get_work_tel(),
                                i.get_mob_tel(),
                                i.get_city(),
                                i.get_filialName(),
                                dl_name_id
                            )

                            print(sql)

                            cursor.execute(sql)
                            conn.commit()

                            cursor.close()



                    except Exception as error:
                        print("Failed to insert record into Laptop table {}".format(error))


                    finally:
                        conn.close()
                        print("Success")

                        try:
                            org_id = re.sub("[^0-9]", "", str(record))
                        except IndexError as error:
                            org_id = 0

                        try:

                            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                    password='postgres', host='localhost')
                            cursor = conn.cursor()
                            sql = 'SELECT * FROM administrators_order_admin_ib WHERE order_num = \'{0}\' AND order_date = \'{1}\''.format(
                                i.get_order_num(),
                                i.get_order_date()
                            )

                            cursor.execute(sql)
                            conn.commit()

                            orderRrecords = cursor.fetchall()

                            if len(orderRrecords) == 0:
                                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                        password='postgres', host='localhost')
                                cursor = conn.cursor()
                                sql = 'SELECT MAX(id) FROM administrators_order_admin_ib'

                                try:
                                    cursor.execute(sql)
                                    conn.commit()

                                    lastOrgId = cursor.fetchall()
                                    lastOrgId = int(re.sub("[^0-9]", "", str(lastOrgId)))
                                    lastOrgId += 1
                                except Exception as e:
                                    lastOrgId = 1
                                sql = 'INSERT INTO public.administrators_order_admin_ib(id, order_num, order_date, order_file, order_administrator_id, order_organization_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', {4}, {5})'.format(
                                    lastOrgId,
                                    i.get_order_num(),
                                    i.get_order_date(),
                                    'mainapp\\admin_ib_orders\\' + str(i.get_gid()) + '.pdf',
                                    adminId,
                                    org_id
                                )

                                print(sql)

                                cursor.execute(sql)
                                conn.commit()

                                cursor.close()

                            else:
                                conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                                        password='postgres', host='localhost')
                                cursor = conn.cursor()
                                orderId = orderRrecords[0][0]

                                sql = 'DELETE FROM administrators_order_admin_ib WHERE id = {0}'.format(orderId)
                                print(sql)
                                try:
                                    cursor.execute(sql)
                                    conn.commit()

                                except Exception as e:
                                    print("Failed to delete from administrators_order_admin_ib table {}".format(error))

                                sql = 'INSERT INTO public.administrators_order_admin_ib(id, order_num, order_date, order_file, order_administrator_id, order_organization_id) VALUES ({0}, \'{1}\', \'{2}\', \'{3}\', {4}, {5})'.format(
                                    orderId,
                                    i.get_order_num(),
                                    i.get_order_date(),
                                    'mainapp\\admin_ib_orders\\' + str(i.get_gid()) + '.pdf',
                                    adminId,
                                    org_id
                                )

                                print(sql)

                                cursor.execute(sql)
                                conn.commit()

                                cursor.close()


                        except Exception as error:
                            print("Failed to insert record into Laptop table {}".format(error))

                        finally:
                            conn.close()
                            print("Success")

def fethDicts():

    orderActivities = []
    orderStatuses = []
    for org in organizations:

        if (org.get_order_activity() != "None") and (org.get_order_activity() not in orderActivities):
            try:
                orderActivities.append(org.get_order_activity().strip())
            except AttributeError as error:
                continue

        if (org.get_order_status != "None") and (org.get_order_status() not in orderStatuses):
            try:
                orderStatuses.append(org.get_order_status().strip())
            except AttributeError as error:
                continue


    conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                            password='postgres', host='localhost')
    cursor = conn.cursor()

    sql = 'SELECT * FROM administrators_orderactivities WHERE activity_name = \'{0}\''.format(
       'Не указан'
    )

    cursor.execute(sql)
    conn.commit()

    activityRrecords = cursor.fetchall()

    if len(activityRrecords) == 0:
        sql = 'INSERT INTO public.administrators_orderactivities(id, activity_name) VALUES (1, \'Не указан\')'

        print(sql)

        cursor.execute(sql)
        conn.commit()


    for orderActivity in orderActivities:
        try:

            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()
            sql = 'SELECT * FROM administrators_orderactivities WHERE activity_name = \'{0}\''.format(
                orderActivity
            )

            cursor.execute(sql)
            conn.commit()

            orderActivitiesBd = cursor.fetchall()

            if len(orderActivitiesBd) == 0:
                sql = 'SELECT MAX(id) FROM administrators_orderactivities'


                cursor.execute(sql)
                conn.commit()

                lastId = cursor.fetchall()
                lastId = int(re.sub("[^0-9]", "", str(lastId)))
                lastId += 1

                sql = 'INSERT INTO public.administrators_orderactivities(id, activity_name) VALUES ({0}, \'{1}\')'.format(
                    lastId,
                    orderActivity
                )

                print(sql)

                cursor.execute(sql)
                conn.commit()





        except Exception as error:
            print("Failed to insert record into Laptop table {}".format(error))

        finally:
            conn.close()
            print("Success")

    conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                            password='postgres', host='localhost')
    cursor = conn.cursor()
    sql = 'SELECT * FROM administrators_orderstatuses WHERE status_name = \'{0}\''.format(
        'Не указан'
    )

    cursor.execute(sql)
    conn.commit()

    orderStatusesBd = cursor.fetchall()
    if len(orderStatusesBd) == 0:
        sql = 'INSERT INTO public.administrators_orderstatuses(id, status_name) VALUES (1, \'Не указан\')'

        print(sql)

        cursor.execute(sql)
        conn.commit()

    for orderStatus in orderStatuses:
        try:


            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                    password='postgres', host='localhost')
            cursor = conn.cursor()
            sql = 'SELECT * FROM administrators_orderstatuses WHERE status_name = \'{0}\''.format(
                orderStatus
            )

            cursor.execute(sql)
            conn.commit()

            orderStatusesBd = cursor.fetchall()

            if len(orderStatusesBd) == 0:
                sql = 'SELECT MAX(id) FROM administrators_orderstatuses'


                cursor.execute(sql)
                conn.commit()

                lastId = cursor.fetchall()
                lastId = int(re.sub("[^0-9]", "", str(lastId)))
                lastId += 1
                sql = 'INSERT INTO public.administrators_orderstatuses(id, status_name) VALUES ({0}, \'{1}\')'.format(
                    lastId,
                    orderStatus
                )

                print(sql)

                cursor.execute(sql)
                conn.commit()



        except Exception as error:
            print("Failed to insert record into Laptop table {}".format(error))

        finally:
            conn.close()
            print("Success")

    dlNameId = 1
    dlList = []
    for dl in dls:

        if (dl.get_dl_name() != "None") and (dl.get_dl_name() not in dlList):
            try:
                dlList.append(dl.get_dl_name().strip())
            except AttributeError as error:
                continue

    for dl in dlList:
        try:

            conn = psycopg2.connect(dbname='OKZ_DB', user='postgres',
                                        password='postgres', host='localhost')
            cursor = conn.cursor()
            sql = 'SELECT * FROM administrators_dlnames WHERE dl_name = \'{0}\''.format(
                   dl
            )

            cursor.execute(sql)
            conn.commit()

            dlnames = cursor.fetchall()

            if len(dlnames) == 0:
                sql = 'SELECT MAX(id) FROM administrators_dlnames'


                cursor.execute(sql)
                conn.commit()

                lastId = cursor.fetchall()
                lastId = int(re.sub("[^0-9]", "", str(lastId)))
                lastId += 1
                sql = 'INSERT INTO public.administrators_dlnames(id, dl_name) VALUES ({0}, \'{1}\')'.format(
                    lastId,
                    dl
                )

                print(sql)

                cursor.execute(sql)
                conn.commit()

                dlNameId += 1

        except Exception as error:
            print("Failed to insert record into Laptop table {}".format(error))

        finally:
            conn.close()
            print("Success")



readOrderBase()
readDlBase()
fethDicts()
fetchOrgDB()
fetchDlDB()




#while True:
