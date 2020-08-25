import os
import re
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook

mypath = 'C:\orders'

orderBaseFile = 'C:\\microservices\\order_base.xlsm'

organizations = []




class Organization:

    def __init__(self, docid,
                 gid,
                 inn,
                 full_organization_name,
                 organization_name,
                 services,
                 order_start_date,
                 order_end_date,
                 order_status,
                 order_activity,
                 document_price,
                 document_price_with_nds,
                 doc_prepare_date,
                 doc_eosdo_date,
                 doc_sending_date,
                 order_activity_date,
                 doc_eosdo_num,
                 doc_eosdo_link,
                 ):
        self.docid = docid
        self.gid = gid
        self.inn = inn
        self.full_organization_name = full_organization_name
        self.organization_name = organization_name
        self.services = services
        self.order_start_date = order_start_date
        self.order_end_date = order_end_date
        self.order_status = order_status
        self.order_activity = order_activity
        self.document_price = document_price
        self.document_price_with_nds = document_price_with_nds
        self.doc_prepare_date = doc_prepare_date
        self.doc_eosdo_date = doc_eosdo_date
        self.doc_sending_date = doc_sending_date
        self.order_activity_date = order_activity_date
        self.doc_eosdo_num = doc_eosdo_num
        self.doc_eosdo_link = doc_eosdo_link

    def set_city(self, city):
        self._city = city

    def set_filial_name(self, filial_name):
        self._filial_name = filial_name

    def set_order_num(self, order_num):
        self._order_num = order_num

    def set_order_date(self, order_date):
        self._order_date = order_date

    def get_gid(self):
        return self.gid

    def get_docid(self):
        return self.docid

    def get_inn(self):
        return self.inn

    def get_full_organization_name(self):
        return self.full_organization_name

    def get_organization_name(self):
        return self.organization_name

    def get_services(self):
        return self.services

    def get_order_start_date(self):
        return self.order_start_date

    def get_order_end_date(self):
        return self.order_end_date

    def get_order_status(self):
        return self.order_status

    def get_order_activity(self):
        return self.order_activity

    def get_document_price(self):
        return self.document_price

    def get_document_price_with_nds(self):
        return self.document_price_with_nds

    def get_doc_prepare_date(self):
        return self.doc_prepare_date

    def get_doc_eosdo_date(self):
        return self.doc_eosdo_date

    def get_doc_sending_date(self):
        return self.doc_sending_date

    def get_order_activity_date(self):
        return self.order_activity_date

    def get_doc_eosdo_num(self):
        return self.doc_eosdo_num

    def get_doc_eosdo_link(self):
        return self.doc_eosdo_link



wb = load_workbook(orderBaseFile, data_only=True)

ws = wb['Договоры']
j = 1
for i in range(1, ws.max_row):
    cell = ws["A" + str(i)]
    if "ID документа" in str(cell.value):
        break
    else:
        j += 1

j += 1

for row in range(j, ws.max_row):

    docid = ws["A" + str(row)].value
    gid = ws["E" + str(row)].value
    inn = ws["F" + str(row)].value
    full_organization_name = ws["H" + str(row)].value
    organization_name = ws["I" + str(row)].value
    services = ws["P" + str(row)].value
    if str(ws["Q" + str(row)].value) == "None":
        order_start_date = '2030-01-01 00:00:00'
    else:
        order_start_date = ws["Q" + str(row)].value

    if str(ws["R" + str(row)].value) == "None":
        order_end_date = '2030-01-01 00:00:00'
    else:
        order_end_date = ws["R" + str(row)].value

    order_status = ws["T" + str(row)].value
    order_activity = ws["U" + str(row)].value
    document_price = ws["AC" + str(row)].value
    document_price_with_nds = ws["AD" + str(row)].value
    if str(ws["AF" + str(row)].value) == "None":
        doc_prepare_date = '2030-01-01 00:00:00'
    else:
        doc_prepare_date = ws["AF" + str(row)].value

    if str(ws["AG" + str(row)].value) == "None":
        doc_eosdo_date = '2030-01-01 00:00:00'
    else:
        doc_eosdo_date = ws["AG" + str(row)].value

    if str(ws["AH" + str(row)].value) == "None":
        doc_sending_date = '2030-01-01 00:00:00'
    else:
        doc_sending_date = ws["AH" + str(row)].value

    if str(ws["AI" + str(row)].value) == "None":
        order_activity_date = '2030-01-01 00:00:00'
    else:
        order_activity_date = ws["AI" + str(row)].value
    doc_eosdo_num = ws["AN" + str(row)].value
    try:
        doc_eosdo_link = ws["AO" + str(row)].hyperlink.target
    except AttributeError:
        oc_eosdo_link = ""

    newOrganization = Organization(docid, gid, inn, full_organization_name, organization_name, services,
                                   order_start_date,
                                   order_end_date,
                                   order_status,
                                   order_activity,
                                   document_price,
                                   document_price_with_nds,
                                   doc_prepare_date,
                                   doc_eosdo_date,
                                   doc_sending_date,
                                   order_activity_date,
                                   doc_eosdo_num,
                                   doc_eosdo_link)

    organizations.append(newOrganization)

onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
regex = re.compile('[^а-яА-Я]')

for i in organizations:
    for j in onlyfiles:
        if i.get_organization_name() in j:

            ext = j[-4:]
            old_file = os.path.join(mypath, j[:-4] + ext)
            new_file = os.path.join(mypath, str(i.get_gid()) + ext)
            try:
                os.rename(old_file, new_file)
            except FileNotFoundError as error:
              continue
