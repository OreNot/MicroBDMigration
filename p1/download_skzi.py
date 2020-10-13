from openpyxl import load_workbook
import psycopg2
from tqdm import tqdm

orderSkziFile = 'C:\\microservices\\skzi_base — new.xlsx'
skzi = []
organization_mas = []

db_config = {'database':'test',
           'user':'postgres',
           'password':'postgres',
           'host':'127.0.0.1'}

conn = psycopg2.connect(**db_config)
cursor = conn.cursor()

'''
class Skzis:

    def __init__(self, skziuser_name, 
                skziserial_name,
                pvmserial_name,
                pechat_number,
                adress,
                szz,
                catalognumber,
                description,
                sertificate,
                skzi_type,
                ais,
                software,
                organization,
                ):
        
        self.skziuser_name = skziuser_name
        self.skziserial_name = skziserial_name
        self.pvmserial_name = pvmserial_name
        self.pechat_number = pechat_numbe
        self.adress = adress
        self.szz = szz
        self.catalognumber = catalognumber
        self.description = description
        self.sertificate = sertificate
        self.skzi_type = skzi_type
        self.ais = ais
        self.software = software
        self.organization = organization
'''

def readSkzis_ais(row_num, ws):

    if (ws.cell(row = row_num, column = 6).value) == None or (ws.cell(row = row_num, column = 6).value).strip() =='':
        ais = 'Null'
    else:
        ais = (ws.cell(row = row_num, column = 6).value).strip()

    _SQL = """SELECT id FROM skzis_ais WHERE ais_name = ('%(ais)s')"""%{'ais':ais}
    cursor.execute(_SQL)
    num_id = cursor.fetchall()

    if num_id == []:

        _SQL = """SELECT COUNT(*) FROM skzis_ais"""
        cursor.execute(_SQL)
        num_id = cursor.fetchall()
        num_id = num_id[0][0] + 1
        
        _SQL = """INSERT INTO skzis_ais (id, ais_name) VALUES ('%(id)s','%(ais_name)s')"""%{'id':num_id, 'ais_name':ais}
        cursor.execute(_SQL)
        conn.commit()

    else:
        num_id = num_id[0][0] 

    return (num_id)


def readSkzis_skzitype(row_num, ws):

    if (ws.cell(row = row_num, column = 4).value) == None or (ws.cell(row = row_num, column = 4).value).strip() =='':
        skzi_type = 'Null'
    else:
        skzi_type = (ws.cell(row = row_num, column = 4).value).strip()

    _SQL = """SELECT id FROM skzis_skzitype WHERE type_name = ('%(type_name)s')"""%{'type_name':skzi_type}
    cursor.execute(_SQL)
    num_id = cursor.fetchall()

    if num_id == []:

        _SQL = """SELECT COUNT(*) FROM skzis_skzitype"""
        cursor.execute(_SQL)
        num_id = cursor.fetchall()
        num_id = num_id[0][0] + 1
        
        _SQL = """INSERT INTO skzis_skzitype (id, type_name) VALUES ('%(id)s','%(type_name)s')"""%{'id':num_id, 'type_name':skzi_type}
        cursor.execute(_SQL)
        conn.commit()
    
    else:
        num_id = num_id[0][0] 

    return (num_id)


def readSkzis_software(row_num, ws):

    if (ws.cell(row = row_num, column = 10).value) == None or (ws.cell(row = row_num, column = 10).value).strip() =='':
        software = 'Null'
    else:
        software = (ws.cell(row = row_num, column = 10).value).strip()

    _SQL = """SELECT id FROM skzis_software WHERE software_name = ('%(software_name)s')"""%{'software_name':software}
    cursor.execute(_SQL)
    num_id = cursor.fetchall()

    if num_id == []:

        _SQL = """SELECT COUNT (*) FROM skzis_software"""
        cursor.execute(_SQL)
        num_id = cursor.fetchall()
        num_id = num_id[0][0] + 1

        _SQL = """INSERT INTO skzis_software (id, software_name) VALUES ('%(id)s','%(software_name)s')"""%{'id':num_id, 'software_name':software}
        cursor.execute(_SQL)
        conn.commit()

    else:
        num_id = num_id[0][0] 
    
    return (num_id)

def readSkzis_organization(row_num, ws):

    if (ws.cell(row = row_num, column = 1).value) == None or (ws.cell(row = row_num, column = 1).value).strip() =='':
        organization = 'Null'
    else:
        organization = (ws.cell(row = row_num, column = 1).value).strip()

    _SQL = """SELECT id FROM administrators_organizations WHERE organization_name = ('%(organization_name)s')"""%{'organization_name':organization}
    cursor.execute(_SQL)
    num_id = cursor.fetchall()

    if num_id == []:
        organization_mas.append(organization) # Список организаций которые нети в ДБ

    else:
        num_id = num_id[0][0] 

    return (num_id) 


def readSkzis():
    wb = load_workbook(orderSkziFile, data_only=True)
    ws = wb['Сх криптозащСКЗИ']

    conn = psycopg2.connect(**db_config)
    cursor = conn.cursor()


    for row_num in range (2, ws.max_row):

        ais_id = readSkzis_ais(row_num,ws)
        skzi_type_id = readSkzis_skzitype(row_num,ws)
        software_id = readSkzis_software(row_num, ws)
        organization_id = readSkzis_organization(row_num, ws)

        if organization_id == []: #Если организации нет, в таблицу skzis_skzis ничего не записываем
            continue

        if (ws.cell(row = row_num, column = 3).value) == None or (ws.cell(row = row_num, column = 3).value).strip() =='':
            skziuser_name = 'Null'
        else:
            skziuser_name = (ws.cell(row = row_num, column = 3).value).strip()

        if (ws.cell(row = row_num, column = 5).value) == None or (ws.cell(row = row_num, column = 5).value).strip() =='':
            skziserial_name = 'Null'
        else:
            skziserial_name = (ws.cell(row = row_num, column = 5).value).strip()

        if (ws.cell(row = row_num, column = 7).value) == None or (ws.cell(row = row_num, column = 7).value).strip() =='':
            pvmserial_name = 'Null'
        else:
            pvmserial_name = (ws.cell(row = row_num, column = 7).value).strip()

        if (ws.cell(row = row_num, column = 8).value) == None or (ws.cell(row = row_num, column = 8).value).strip() =='':
            pechat_number = 'Null'
        else:
            pechat_number = (ws.cell(row = row_num, column = 8).value).strip()

        if (ws.cell(row = row_num, column = 9).value) == None or (ws.cell(row = row_num, column = 9).value).strip() =='':
            adress = 'Null'
        else:
            adress = (ws.cell(row = row_num, column = 9).value).strip()

        if (ws.cell(row = row_num, column = 11).value) == None or (ws.cell(row = row_num, column = 11).value).strip() =='':
            szz = 'Null'
        else:
            szz = (ws.cell(row = row_num, column = 11).value).strip()

        if (ws.cell(row = row_num, column = 12).value) == None or ((str(ws.cell(row = row_num, column = 12).value))).strip() =='':
            catalognumber = 'Null'
        else:
            catalognumber = (str(ws.cell(row = row_num, column = 12).value)).strip()

        if (ws.cell(row = row_num, column = 13).value) == None or (ws.cell(row = row_num, column = 13).value).strip() =='':
            description = 'Null'
        else:
            description = (ws.cell(row = row_num, column = 13).value).strip()

        if (ws.cell(row = row_num, column = 14).value) == None or (ws.cell(row = row_num, column = 14).value).strip() =='':
            sertificate = 'Null'
        else:
            sertificate = (ws.cell(row = row_num, column = 14).value).strip()


        _SQL = """INSERT INTO skzis_skzis (skziuser_name, skziserial_name, pvmserial_name, pechat_number, adress, szz, catalognumber, description, sertificate, skzi_type_id, ais_id, software_id, organization_id) VALUES
        ('%(skziuser_name)s', '%(skziserial_name)s', '%(pvmserial_name)s', '%(pechat_number)s', '%(adress)s', '%(szz)s', '%(catalognumber)s', '%(description)s', '%(sertificate)s', '%(skzi_type_id)s', '%(ais_id)s', '%(software_id)s', '%(organization_id)s')
        """%{'skziuser_name':skziuser_name, 'skziserial_name':skziserial_name, 'pvmserial_name':pvmserial_name, 'pechat_number':pechat_number, 'adress':adress, 'szz':szz, 'catalognumber':catalognumber, 'description':description, 'sertificate':sertificate, 'skzi_type_id':skzi_type_id, 'ais_id':ais_id, 'software_id':software_id, 'organization_id':organization_id}

        cursor.execute(_SQL)
        conn.commit()
    conn.close()

readSkzis()

for i in range (len(organization_mas)): # выводим на печать список не обработанных организаций
    print (organization_mas[i])

print (len(organization_mas)) # выводим на печать количество не обработанных строк